import io, re
from datetime import datetime
import streamlit as st
from openpyxl import load_workbook
from pdfminer.high_level import extract_text
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

# Optional flatten backend
try:
    import fitz  # PyMuPDF
    from PIL import Image
    HAS_FITZ = True
except Exception:
    HAS_FITZ = False

# ---------- Common helpers ----------
SIDE_MARGIN_MM = 2
TOP_MARGIN_MM = 4
STAMP_BOTTOM_MM = 12
INTER_GAP_MM = 1
BASE_CROP_L = 6
BASE_CROP_R = 6
BASE_CROP_T = 8
BASE_CROP_B = 8
LOW_TEXT_LINES = 4
SHORT_TEXT_CHARS = 80
EXTRA_CROP_LR = 14
EXTRA_CROP_T  = 18
EXTRA_CROP_B  = 28
RASTER_SCALE = 2.0  # used only when HAS_FITZ=True

def strip_diacritics(s: str) -> str:
    import unicodedata
    if s is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFKD", str(s)) if ord(c) < 128)

def read_excel_lookup(file_like):
    wb = load_workbook(file_like, data_only=True)
    ws = wb.active
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[str(v).strip().lower()] = col
    z_col = headers.get("zlecenie")
    ilo_col = headers.get("ilośc palet") or headers.get("ilosc palet") or headers.get("ilość palet")
    pr_col = headers.get("przewoźnik") or headers.get("przewoznik")
    if not z_col or not ilo_col or not pr_col:
        raise ValueError("Excel musi mieć kolumny: ZLECENIE, ilość palet, przewoźnik (nagłówki w 1. wierszu).")
    lookup = {}
    all_nums = set()
    for row in range(2, ws.max_row + 1):
        z = ws.cell(row=row, column=z_col).value
        il = ws.cell(row=row, column=ilo_col).value
        pr = ws.cell(row=row, column=pr_col).value
        z = "" if z is None else str(z).strip()
        il = "" if il is None else str(il).strip()
        pr = "" if pr is None else str(pr).strip()
        parts = [p.strip() for p in re.split(r"[+;,/\\s]+", z) if p.strip()]
        for p in parts:
            p2 = "".join(ch for ch in p if ch.isdigit())
            if p2.isdigit():
                all_nums.add(p2)
                lookup[p2] = (z, il, pr)
    return lookup, all_nums

NBSP = "\\u00A0"; NNBSP = "\\u202F"; THINSP = "\\u2009"
def normalize_digits(s: str) -> str:
    import re
    return re.sub(r"[\\s\\-{}{}{}]".format(NBSP, NNBSP, THINSP), "", s)

def extract_candidates(text: str):
    import re
    normal = re.findall(r"\\b\\d{4,8}\\b", text)
    fancy = re.findall(r"(?<!\\d)(?:\\d[\\s\\u00A0\\u202F\\u2009\\-]?){4,9}(?!\\d)", text)
    fancy = [normalize_digits(s) for s in fancy]
    so = [normalize_digits(m.group(1)) for m in re.finditer(r"Sales\\s*[\\r\\n ]*Order[\\s:]*([0-9\\s\\u00A0\\u202F\\u2009\\-]{4,12})", text, flags=re.I)]
    cands = normal + fancy + so
    cands = [c for c in cands if c.isdigit() and 4 <= len(c) <= 8]
    out, seen = [], set()
    for c in cands:
        if c not in seen:
            out.append(c); seen.add(c)
    return out

def adaptive_crop_extra(text: str):
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    sparse = (len(lines) <= LOW_TEXT_LINES) or (len((text or "")) < SHORT_TEXT_CHARS)
    if sparse:
        return (EXTRA_CROP_LR*mm, EXTRA_CROP_LR*mm, EXTRA_CROP_T*mm, EXTRA_CROP_B*mm)
    return (0,0,0,0)

def make_stamp(canvas_obj, W, header, footer):
    try:
        canvas_obj.setFont("Helvetica-Bold", 12)
    except Exception:
        canvas_obj.setFont("Helvetica", 12)
    canvas_obj.drawRightString(W - 8*mm, 8*mm + 12 + 1, header)
    if footer:
        canvas_obj.drawRightString(W - 8*mm, 8*mm, footer)

# ---------- Flatten path (if fitz available) ----------
def rasterize_and_crop(page, cl, cr, ct, cb, scale):
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    left   = int(cl * scale)
    right  = int((page.rect.width - cr) * scale)
    top    = int(ct * scale)
    bottom = int((page.rect.height - cb) * scale)
    left = max(0, min(left, img.width-1)); right = max(left+1, min(right, img.width))
    top = max(0, min(top, img.height-1)); bottom = max(top+1, min(bottom, img.height))
    return img.crop((left, top, right, bottom))

def engine_flatten(pdf_bytes, xlsx_bytes, max_per_sheet):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    lookup, excel_numbers = read_excel_lookup(io.BytesIO(xlsx_bytes))

    groups, page_meta, page_text_cache = {}, {}, {}
    for i in range(len(doc)):
        page_text = extract_text(io.BytesIO(pdf_bytes), page_numbers=[i]) or ""
        page_text_cache[i] = page_text
        cands = extract_candidates(page_text)
        picked = next((n for n in cands if n in excel_numbers), None)
        mapped = lookup.get(picked) if picked else None
        if mapped:
            z_full, il, pr = mapped
            key = z_full
            header = ("ZLECENIA (laczone): {}".format(strip_diacritics(z_full))
                      if "+" in z_full else "ZLECENIE: {}".format(strip_diacritics(z_full)))
            footer = "ilosc palet: {} | przewoznik: {}".format(strip_diacritics(il), strip_diacritics(pr))
        elif picked:
            key = picked
            header = "ZLECENIE: {}".format(picked)
            footer = "(brak danych w Excelu)"
        else:
            key = "_NO_ORDER_{}".format(i+1)
            header = "(nie znaleziono numeru zlecenia na tej stronie)"
            footer = ""
        groups.setdefault(key, []).append(i)
        page_meta[i] = (header, footer)

    def key_sort(k: str):
        import re
        nums = [int(x) for x in re.findall(r"\\d+", k)]
        return (min(nums) if nums else 10**9, k)
    ordered_keys = sorted(groups.keys(), key=key_sort)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    margin_x = SIDE_MARGIN_MM * mm
    top_margin = TOP_MARGIN_MM * mm
    bottom_for_stamp = STAMP_BOTTOM_MM * mm
    gap = INTER_GAP_MM * mm
    avail_w = W - 2 * margin_x
    avail_h = H - top_margin - bottom_for_stamp
    base_crop_l = BASE_CROP_L * mm
    base_crop_r = BASE_CROP_R * mm
    base_crop_t = BASE_CROP_T * mm
    base_crop_b = BASE_CROP_B * mm

    for gkey in ordered_keys:
        idxs = groups[gkey]
        for start in range(0, len(idxs), max_per_sheet):
            batch = idxs[start:start+max_per_sheet]
            items, total_h = [], 0.0
            for idx in batch:
                page = doc[idx]
                sw = float(page.rect.width); sh = float(page.rect.height)
                ex_l, ex_r, ex_t, ex_b = adaptive_crop_extra(page_text_cache[idx])
                cl = base_crop_l + ex_l; cr = base_crop_r + ex_r
                ct = base_crop_t + ex_t; cb = base_crop_b + ex_b
                cw = max(10.0, sw - cl - cr); ch = max(10.0, sh - ct - cb)
                s  = avail_w / cw; dh = s * ch
                items.append((idx, cl, cr, ct, cb, s, dh))
                total_h += dh
            total_h += gap * max(0, len(batch)-1)
            down = min(1.0, avail_h / total_h) if total_h > 0 else 1.0

            y = H - top_margin
            for (idx, cl, cr, ct, cb, s, dh) in items:
                s *= down; dh *= down
                cropped = rasterize_and_crop(doc[idx], cl, cr, ct, cb, RASTER_SCALE)
                img_bytes = io.BytesIO(); cropped.save(img_bytes, format="PNG"); img_bytes.seek(0)
                x = margin_x; y2 = y - dh
                c.drawImage(ImageReader(img_bytes), x, y2, width=avail_w, height=dh, preserveAspectRatio=False, mask='auto')
                y = y2 - gap

            make_stamp(c, W, *page_meta[batch[0]])
            c.showPage()
    c.save()
    return buf.getvalue()

# ---------- Sanitize path (fallback without fitz) ----------
def engine_sanitize(pdf_bytes, xlsx_bytes, max_per_sheet):
    from PyPDF2 import PdfReader, PdfWriter, Transformation
    from PyPDF2._page import PageObject
    lookup, excel_numbers = read_excel_lookup(io.BytesIO(xlsx_bytes))
    reader = PdfReader(io.BytesIO(pdf_bytes))

    groups, page_meta, page_text_cache = {}, {}, {}
    for i, _ in enumerate(reader.pages):
        page_text = extract_text(io.BytesIO(pdf_bytes), page_numbers=[i]) or ""
        page_text_cache[i] = page_text
        cands = extract_candidates(page_text)
        picked = next((n for n in cands if n in excel_numbers), None)
        mapped = lookup.get(picked) if picked else None
        if mapped:
            z_full, il, pr = mapped
            key = z_full
            header = ("ZLECENIA (laczone): {}".format(strip_diacritics(z_full))
                      if "+" in z_full else "ZLECENIE: {}".format(strip_diacritics(z_full)))
            footer = "ilosc palet: {} | przewoznik: {}".format(strip_diacritics(il), strip_diacritics(pr))
        elif picked:
            key = picked; header = "ZLECENIE: {}".format(picked); footer = "(brak danych w Excelu)"
        else:
            key = "_NO_ORDER_{}".format(i+1); header = "(nie znaleziono numeru zlecenia na tej stronie)"; footer = ""
        groups.setdefault(key, []).append(i); page_meta[i] = (header, footer)

    def key_sort(k: str):
        import re
        nums = [int(x) for x in re.findall(r"\\d+", k)]; return (min(nums) if nums else 10**9, k)
    ordered_keys = sorted(groups.keys(), key=key_sort)

    from reportlab.lib.pagesizes import A4
    W, H = A4
    margin_x = SIDE_MARGIN_MM * mm
    top_margin = TOP_MARGIN_MM * mm
    bot_stamp  = STAMP_BOTTOM_MM * mm
    gap = INTER_GAP_MM * mm
    avail_w = W - 2 * margin_x
    avail_h = H - top_margin - bot_stamp
    base_crop_l = BASE_CROP_L * mm
    base_crop_r = BASE_CROP_R * mm
    base_crop_t = BASE_CROP_T * mm
    base_crop_b = BASE_CROP_B * mm

    writer = PdfWriter(); writer.add_metadata({"/Producer": "Kersia PDF Stamper Web v1.3 (sanitize)"})
    for gkey in ordered_keys:
        idxs = groups[gkey]
        for start in range(0, len(idxs), max_per_sheet):
            batch = idxs[start:start+max_per_sheet]
            items, total_h = [], 0.0
            for idx in batch:
                sw = float(reader.pages[idx].mediabox.width); sh = float(reader.pages[idx].mediabox.height)
                ex_l, ex_r, ex_t, ex_b = adaptive_crop_extra(page_text_cache[idx])
                cl = base_crop_l + ex_l; cr = base_crop_r + ex_r
                ct = base_crop_t + ex_t; cb = base_crop_b + ex_b
                cw = max(10.0, sw - cl - cr); ch = max(10.0, sh - ct - cb)
                s  = avail_w / cw; dh = s * ch
                items.append((idx, cl, cr, ct, cb, s, dh)); total_h += dh
            total_h += gap * max(0, len(batch)-1)
            down = min(1.0, avail_h / total_h) if total_h > 0 else 1.0

            writer.add_blank_page(width=W, height=H); base_page = writer.pages[-1]
            y = H - top_margin
            for (idx, cl, cr, ct, cb, s, dh) in items:
                s *= down; dh *= down
                x = margin_x - s * cl; y2 = y - dh
                tmp = PageObject.create_blank_page(width=W, height=H)
                tmp.merge_page(reader.pages[idx])
                T = (Transformation().translate(-cl, -cb).scale(s, s).translate(x, y2))
                tmp.add_transformation(T); base_page.merge_page(tmp)
                y = y2 - gap

            from PyPDF2 import PdfReader as _R
            ov = _R(io.BytesIO(_make_overlay(W, H, *page_meta[batch[0]])))
            base_page.merge_page(ov.pages[0])

    buf = io.BytesIO(); writer.write(buf); buf.seek(0)
    # dodatkowe przepisanie
    r = PdfReader(buf, strict=False); w = PdfWriter()
    for p in r.pages: w.add_page(p)
    out = io.BytesIO(); w.write(out); return out.getvalue()

def _make_overlay(width, height, header, footer, font_size=12, margin_mm=8):
    buf = io.BytesIO(); c = canvas.Canvas(buf, pagesize=(width, height))
    try: c.setFont("Helvetica-Bold", font_size)
    except Exception: c.setFont("Helvetica", font_size)
    m = margin_mm * mm; c.drawRightString(width - m, m + font_size + 1, header)
    if footer: c.drawRightString(width - m, m, footer)
    c.save(); return buf.getvalue()

def annotate_pdf_web(pdf_bytes, xlsx_bytes, max_per_sheet):
    if HAS_FITZ:
        return engine_flatten(pdf_bytes, xlsx_bytes, max_per_sheet)
    else:
        return engine_sanitize(pdf_bytes, xlsx_bytes, max_per_sheet)

# ---------- UI ----------
st.set_page_config(page_title="Kersia PDF Stamper v1.4c", page_icon="🧰", layout="centered")
st.title("Kersia — PDF Stamper (auto-detekcja silnika)")
if not 'notified' in st.session_state:
    if not HAS_FITZ:
        st.warning("Działa tryb zapasowy (bez PyMuPDF). Aby mieć 100% zgodność z Adobe, dodaj do repo **runtime.txt** z `python-3.12.0` i zależność `pymupdf` w requirements.")
    st.session_state['notified'] = True

excel_file = st.file_uploader("Plik Excel:", type=["xlsx", "xlsm", "xls"])
pdf_file = st.file_uploader("Plik PDF:", type=["pdf"])
max_per_sheet = st.slider("Maks. stron na kartkę", 1, 6, 3, 1)

if st.button("GENERUJ PDF", type="primary", disabled=not (excel_file and pdf_file)):
    try:
        data = annotate_pdf_web(pdf_file.read(), excel_file.read(), max_per_sheet)
        fname = "zlecenia_{}.pdf".format(datetime.now().strftime('%Y%m%d'))
        st.success("Gotowe! Pobierz poniżej.")
        st.download_button("Pobierz wynik", data=data, file_name=fname, mime="application/pdf")
    except Exception as e:
        st.error("Błąd: {}".format(repr(e)))
